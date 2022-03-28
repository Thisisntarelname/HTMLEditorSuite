#!/usr/bin/env python
#This takes the xlsx file from /uploads, edits the /Data pages to reflect upadted information from xlsx file. 
#pip3 install beautifulsoup4
#pip3 install lxml
import os
import openpyxl
import glob
import json
from bs4 import BeautifulSoup
import lxml


filenames = []
teachernames = []


for file in glob.glob("uploads/*.xlsx"):
  path = (file)#Needs to open the one file in 'uploads' folder
  print(path)
  delete = path


wb_obj = openpyxl.load_workbook(path)
sheet_obj = wb_obj.active
ws = sheet_obj
 
sheet_object = wb_obj.active
max_rows = sheet_object.max_row#finding how many rows there are

os.chdir('Data')
i = 2#if i=1, it would use "21/22" (the year) as well
while i <= max_rows:

  rooms = []
  subjects = []
  classes = []
  #The three empties every time it runs, clearing it for the next name

  cell_obj = sheet_obj.cell(row = i, column = 1)
  if (cell_obj.value is None) or (cell_obj.value == "GYM") or (cell_obj.value == "DANCE"):#filters out empty, "DANCE", and "GYM" results
    a = 1
  else:
    txt = cell_obj.value
    x = "RM " in txt
    print(txt)
    
    if (x == True):#filters out room number results


      
      a = 1 
    else:
        #This is where we get the room number, it tales the cell below it and gets rif of "RM ", and then appends it to 'rooms'-----
        room_number = sheet_obj.cell(row = i+1, column = 1)
        room_number = room_number.value
        rooms.append(room_number.replace("RM ", ""))
        #------
        #This is where we get the classes-------
        ii = 1
        while ii <=10:#10 because that's how many columns there are total
          class_test = sheet_obj.cell(row = i, column = ii)
          class_test = class_test.value
          #Now 'class_test' is the value of a cell (i, ii) being the coordinates. We can check to see if it's in 'class' already. If it is, we can ignore it. If it's not we know it's a new class and will add it. But first, we need to check if it's value is "None" and get rid of the "\n(teacher)"
          if class_test != None:
            if "\n" in class_test:
              #delete the "\n" and the parantheses, including everything inside the parantheses
              v = 0
              for v in range(0, len(class_test)):
                if class_test[v] == "(":
                  start = v
                if class_test[v] == ")":
                  end = v
              vv = start
              class_test_list = []
              class_test_list[:0]=class_test
              while vv<=end:
                del class_test_list[start]
                vv = vv+1

              class_test_list.remove("\n")
              
              class_test = "" 
    
              for ele in class_test_list: 
                class_test += ele 
            a = 1
                    
            
            if class_test not in classes and class_test != None:
              class_test = str(class_test)
              classes.append(class_test)
          ii = ii+1

        #deleting the teachers' name from the forst index, somehow that got in there
        classes.pop(0)
        if "PREP" in classes:
          classes.remove("PREP")
        print(classes)
      
      
      
      



        #-----------------------------------
           
        searchThis = cell_obj.value
        teacherName = searchThis
        searchThis = searchThis.replace(" ", "")

        teachernames.append(teacherName)
        filenames.append(searchThis)

        #This is where we open the html file and edit it
        os.chdir("..")
        with open('TeacherTemplate', 'r') as f:
      
          contents = f.read()
          
      
          soup = BeautifulSoup(contents, 'lxml')

          
        os.chdir('Data')

        #---------Subejcts
        tag = (soup.find(id='subjects'))#Finds h3 with id: subjects
        subjects_master_list = [""]
        build_subjects = subjects[0]#Starting the loop off with first item
        r = 1
        while r<len(subjects):#couldn't get a for loop to work
          build_subjects = build_subjects+", "+subjects[r]#Concated each item to variable passed form last loop through
          r = r+1
    
        subject_string = '<h3 id="subject">Subjects: '+build_subjects+'</h3>'
        tag.replace_with(subject_string)
        #---------Classes
    
        tag = (soup.find(id='classes'))
        build_classes = classes[0]#only adding one class---------------------------------------------------------------------------------------------
        c = 1
        while c<len(classes):
          build_classes = build_classes+", "+classes[c]
          c = i+1
    
        class_string = '<h3 id="classes">Classes: '+build_classes+'</h3>'
        tag.replace_with(class_string)
        #---------Rooms
    
        tag = (soup.find(id='rooms'))
        #rooms = ["201", "302"]#Keep room numbers in string, integers would prpbably require some tweaking of the code
        build_rooms = rooms[0]
        g = 1
        while i<len(rooms):
          build_rooms = build_rooms+", "+rooms[g]
          g = i+1
    
        room_string = '<h3 id="rooms">Rooms: '+build_rooms+'</h3>'
        tag.replace_with(room_string)
      
      
        with open(searchThis, "w") as f_output:
    
          f_output.write(soup.prettify(formatter=None)) 
        
    
      






  i = i+1
  

print("filename", filenames)
print("teachernames", teachernames)


